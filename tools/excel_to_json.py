#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Extract Excel data from سوق الخضار ERP file → JSON for Flutter import.
Usage:  python excel_to_json.py  [output.json]
"""

import sys, io, json, uuid, re
from datetime import datetime
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import openpyxl

EXCEL_PATH = r'G:\flutter\Desktop\سوق_الخضار_ERP_محدث.xlsx'

def parse_date(v):
    if v is None or str(v).strip() == '':
        return None
    v = str(v).strip()
    for fmt in ['%Y-%m-%d %H:%M:%S', '%Y-%m-%d', '%Y/%m/%d']:
        try:
            return datetime.strptime(v, fmt).isoformat()
        except ValueError:
            continue
    return v

def parse_num(v):
    if v is None or str(v).strip() == '':
        return None
    v = str(v).strip().replace(',', '')
    try:
        return float(v)
    except ValueError:
        return None

def clean_str(v):
    if v is None:
        return ''
    s = str(v).strip()
    return s

def detect_pricing_mode(row):
    """Determine pricing mode based on row data."""
    cost_price = parse_num(row[5])  # سعر الطماطم للبرنيكة
    total_cost = parse_num(row[7])  # إجمالي التكلفة
    if cost_price is not None and cost_price > 0:
        return 'per_unit'
    if total_cost is None or total_cost == 0:
        return 'commission'
    return 'lump_sum'

def extract_suppliers_shipments(ws):
    """Extract supplier shipment data."""
    suppliers = {}
    shipments = []
    seen_suppliers = set()

    for i in range(4, ws.max_row + 1):
        row = [ws.cell(row=i, column=c).value for c in range(1, 12)]
        shipment_num = clean_str(row[3])
        if not shipment_num:
            continue

        supplier_name = clean_str(row[2])
        if not supplier_name:
            # Some rows have empty supplier name - use shipment number
            supplier_name = f'مورد {shipment_num}'

        # Create supplier entry
        if supplier_name not in seen_suppliers:
            seen_suppliers.add(supplier_name)
            supplier_id = str(uuid.uuid4())
            suppliers[supplier_name] = {
                'id': supplier_id,
                'name': supplier_name,
                'phone': '',
                'address': '',
                'status': 'Active',
            }

        supplier_id = suppliers[supplier_name]['id']
        date = parse_date(row[1])
        total_barnika = parse_num(row[4]) or 0
        cost_per = parse_num(row[5])
        container_price = parse_num(row[6])
        total_cost = parse_num(row[7]) or 0
        sold = parse_num(row[8]) or 0
        remaining = parse_num(row[9]) or 0
        notes = clean_str(row[10])
        pricing_mode = detect_pricing_mode(row)

        shipment = {
            'id': str(uuid.uuid4()),
            'supplierId': supplier_id,
            'supplierName': supplier_name,
            'shipmentNumber': shipment_num,
            'date': date,
            'pricingMode': pricing_mode,
            'totalBarnikaCount': int(total_barnika),
            'costPricePerBarnika': cost_per if pricing_mode == 'per_unit' else None,
            'lumpSumCost': total_cost if pricing_mode == 'lump_sum' else None,
            'commissionPercentage': None,  # Not in Excel, default
            'containerPrice': container_price,
            'totalCost': total_cost,
            'barnikaSoldCount': int(sold),
            'barnikaRemainingCount': int(remaining),
            'notes': notes,
        }
        shipments.append(shipment)

    return list(suppliers.values()), shipments

def extract_customers_sales(ws, shipments_by_number):
    """Extract customer and sales data."""
    customers = {}
    sales = []
    seen_customers = set()
    empty_barnika_records = []

    for i in range(4, ws.max_row + 1):
        row = [ws.cell(row=i, column=c).value for c in range(1, 17)]
        customer_name = clean_str(row[2])
        if not customer_name:
            continue

        # Create customer entry
        if customer_name not in seen_customers:
            seen_customers.add(customer_name)
            customer_id = str(uuid.uuid4())
            customers[customer_name] = {
                'id': customer_id,
                'name': customer_name,
                'phone': '',
                'address': '',
                'status': 'Active',
                'type': clean_str(row[3]),  # تاجر / مستهلك
            }

        customer_id = customers[customer_name]['id']
        date = parse_date(row[1])
        shipment_num = clean_str(row[4])
        qty = parse_num(row[5])
        price_per = parse_num(row[6])
        total_amount = parse_num(row[7]) or 0
        paid = parse_num(row[8]) or 0
        discount = parse_num(row[9]) or 0
        remaining = parse_num(row[10]) or 0
        due_date = parse_date(row[11])
        payment_status = clean_str(row[12])
        barnika_taken = parse_num(row[13]) or 0
        barnika_remaining_customer = parse_num(row[14]) or 0
        notes = clean_str(row[15])

        # Map to shipment
        shipment_info = shipments_by_number.get(shipment_num)
        shipment_id = shipment_info['id'] if shipment_info else None
        pricing_mode = shipment_info['pricingMode'] if shipment_info else None
        commission_pct = shipment_info.get('commissionPercentage') if shipment_info else None

        # Determine payment method and status
        if payment_status == 'مسدد بالكامل':
            pay_method = 'cash'
            inv_status = 'paid'
            cash_amount = total_amount
            credit_amount = 0
        elif payment_status == 'آجل بالكامل':
            pay_method = 'credit'
            inv_status = 'pending'
            cash_amount = 0
            credit_amount = total_amount
        elif payment_status == 'مسدد جزئياً':
            pay_method = 'cash'
            inv_status = 'partial'
            cash_amount = paid
            credit_amount = remaining
        else:
            pay_method = 'cash' if paid == total_amount else 'credit'
            inv_status = 'paid' if paid >= total_amount else ('partial' if paid > 0 else 'pending')
            cash_amount = paid
            credit_amount = remaining

        sale = {
            'customerId': customer_id,
            'customerName': customer_name,
            'date': date,
            'shipmentNumber': shipment_num,
            'shipmentId': shipment_id,
            'quantity': int(qty) if qty else 0,
            'pricePerUnit': price_per,
            'totalAmount': total_amount,
            'paidAmount': paid,
            'cashAmount': cash_amount,
            'creditAmount': credit_amount,
            'discount': discount,
            'remainingAmount': remaining,
            'dueDate': due_date,
            'paymentStatus': payment_status,
            'status': inv_status,
            'paymentMethod': pay_method,
            'pricingMode': pricing_mode,
            'commissionPercentage': commission_pct,
            'notes': notes,
        }
        sales.append(sale)

        # Empty barnika tracking
        if barnika_taken > 0:
            empty_barnika_records.append({
                'customerId': customer_id,
                'customerName': customer_name,
                'dateOut': date,
                'quantityOut': int(barnika_taken),
                'quantityReturned': int(barnika_taken - barnika_remaining_customer),
                'dateReturned': None,
                'status': 'returned' if barnika_remaining_customer == 0 else 'partial',
                'notes': notes,
            })

    return list(customers.values()), sales, empty_barnika_records

def extract_expenses(ws):
    """Extract expense data."""
    expenses = []
    for i in range(4, ws.max_row + 1):
        row = [ws.cell(row=i, column=c).value for c in range(1, 7)]
        seq = clean_str(row[0])
        if not seq or not seq.isdigit():
            continue
        expenses.append({
            'date': parse_date(row[1]),
            'category': clean_str(row[2]),
            'description': clean_str(row[3]),
            'amount': parse_num(row[4]) or 0,
            'notes': clean_str(row[5]),
        })
    return expenses

def extract_collections(ws):
    """Extract collection/payment data against credit sales."""
    collections = []
    for i in range(2, ws.max_row + 1):
        row = [ws.cell(row=i, column=c).value for c in range(1, 6)]
        seq = clean_str(row[0])
        if not seq or not seq.isdigit():
            continue
        collections.append({
            'date': parse_date(row[1]),
            'customerName': clean_str(row[2]),
            'amount': parse_num(row[3]) or 0,
            'notes': clean_str(row[4]),
        })
    return collections


def main():
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)

    # Extract shipments + suppliers
    suppliers, shipments = extract_suppliers_shipments(wb['موردين'])
    print(f'Found {len(suppliers)} suppliers, {len(shipments)} shipments')

    # Build shipment lookup by number
    shipments_by_number = {}
    for s in shipments:
        shipments_by_number[s['shipmentNumber']] = s
    # Also add by supplierName for matching
    for s in shipments:
        key = f"{s['supplierName']}-{s['shipmentNumber']}"
        shipments_by_number[key] = s

    # Extract customers + sales
    customers, sales, empty_barnika = extract_customers_sales(wb['زباين'], shipments_by_number)
    print(f'Found {len(customers)} customers, {len(sales)} sales, {len(empty_barnika)} empty barnika records')

    # Extract expenses
    expenses = extract_expenses(wb['مصروفات'])
    print(f'Found {len(expenses)} expense entries')

    # Extract collections
    collections = extract_collections(wb['تحصيلات'])
    print(f'Found {len(collections)} collection entries')

    # Build output
    output = {
        'exportDate': datetime.now().isoformat(),
        'source': 'سوق الخضار ERP Excel',
        'suppliers': suppliers,
        'customers': customers,
        'shipments': shipments,
        'sales': sales,
        'expenses': expenses,
        'collections': collections,
        'emptyBarnikaRecords': empty_barnika,
    }

    output_path = sys.argv[1] if len(sys.argv) > 1 else r'G:\development\POS-Offline-Desktop-main\data\vegetable_market_import.json'
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(output, f, ensure_ascii=False, indent=2)

    print(f'\nOutput written to: {output_path}')
    print(f'File size: {len(json.dumps(output, ensure_ascii=False))} chars')

if __name__ == '__main__':
    main()
